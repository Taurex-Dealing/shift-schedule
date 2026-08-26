"""
Build Sep26 from the Excel master and add it to the live shift-schedule tool.

Staffing changes for Sep26 ONLY (prior months untouched - HR backdating shift pay):
  - Nikolay Shumanov, Adam Taylor, Zahid removed (left the company)
  - Andrey added (took over from Zahid, same rota slot)
  - 12 staff total

Also patches BOTH HTML files with the new month plumbing:
  month button + currentMonth default + getMonthInfo entry + embedded data + BUILD_TS.

Run modes:
  python build_september.py            -> DRY RUN: extract + validate, no writes
  python build_september.py --write    -> write data.json + v.txt + patch both HTML files
"""
import json, time, re, os, sys
import openpyxl

REPO = r'C:\Users\SeanConway\td-schedule'
XLSX = r'C:\Global Management Folder\Compenstaions calculations 2026 new.xlsx'
DJ   = os.path.join(REPO, 'data.json')

# August sheet row layout: blanks where Nikolay (offset 5) and Adam Taylor (offset 10)
# used to be; Andrey now in Zahid's old slot (offset 14).
EXCEL_ORDER = ["Hristo","Victor","Stanimir","Ognyan","Petar",None,
               "Matthew","Win","Sarang",None,"Adam Nor","Afnan",
               None,"Andrey","Sean"]
OUTPUT_ORDER = [n for n in EXCEL_ORDER if n]

# hdr row -> (label, expected August dates Mon..Sun; None = Oct spillover)
WEEKS = [
    (3,  "Sep 1-6",   [None, 1, 2, 3, 4, 5, 6]),
    (19, "Sep 7-13",  [7, 8, 9, 10, 11, 12, 13]),
    (35, "Sep 14-20", [14, 15, 16, 17, 18, 19, 20]),
    (51, "Sep 21-27", [21, 22, 23, 24, 25, 26, 27]),
    (67, "Sep 28-30", [28, 29, 30, None, None, None, None]),
]

ROLE = {
    "Morning (7 MYT)": "night",
    "Morning Shift":   "morning",
    "Evening Shift":   "afternoon",
}

def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s != '' else None

def extract():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb['September26']
    weeks, weekend_shifts, problems = [], [], []

    for hdr, label, dates in WEEKS:
        # sanity: Excel header day numbers must match expected August dates
        for i, c in enumerate(range(2, 9)):
            v = ws.cell(hdr, c).value
            v = int(v) if isinstance(v, (int, float)) else None
            if dates[i] is not None and v != dates[i]:
                raise SystemExit(f"Header row {hdr} col {c}: expected {dates[i]}, Excel has {v}")
        staff = []
        wknd = {5: {}, 6: {}}
        for off in range(1, 16):
            r = hdr + off
            name = clean(ws.cell(r, 1).value)
            if name != EXCEL_ORDER[off-1]:
                raise SystemExit(f"Row {r}: expected {EXCEL_ORDER[off-1]!r}, got {name!r}")
            if name is None:
                continue
            shifts = []
            for i, c in enumerate(range(2, 9)):
                if dates[i] is None:
                    shifts.append(None)  # Oct spillover -> null regardless of cell content
                else:
                    shifts.append(clean(ws.cell(r, c).value))
            for idx in (5, 6):
                if dates[idx] is not None and shifts[idx] is not None:
                    wknd[idx][name] = shifts[idx]
            staff.append({"name": name, "shifts": shifts})
        weeks.append({"week": label, "dates": dates, "staff": staff})
        for idx in (5, 6):
            if dates[idx] is None:
                continue
            slot = {"night": "", "morning": "", "afternoon": ""}
            for nm, sh in wknd[idx].items():
                role = ROLE.get(sh)
                if role is None:
                    problems.append(f"{label} {dates[idx]:02d}.09: unmapped weekend shift {sh!r} for {nm}")
                    continue
                if slot[role]:
                    problems.append(f"{label} {dates[idx]:02d}.09: role {role} double-booked ({slot[role]} & {nm})")
                slot[role] = nm
            weekend_shifts.append({"date": f"{dates[idx]:02d}.09", **slot})
    return {"weeks": weeks, "staff_order": OUTPUT_ORDER, "weekend_shifts": weekend_shifts}, problems

def validate(new):
    print("=== VALIDATION ===")
    ok = True
    sp = 0
    for w in new['weeks']:
        dts = [d for d in w['dates'] if d is not None]
        if dts != sorted(dts):
            print(f"  !! spillover/order issue in {w['week']}: {w['dates']}")
            sp += 1
    print(f"  spillover check: {'PASS' if sp==0 else 'FAIL'}")
    ok &= sp == 0
    n = len(new['staff_order'])
    print(f"  staff_order count: {n} (expect 12): {'PASS' if n==12 else 'FAIL'}")
    ok &= n == 12
    wk = len(new['weekend_shifts'])
    print(f"  weekend_shifts entries: {wk} (expect 8): {'PASS' if wk==8 else 'FAIL'}")
    ok &= wk == 8
    nights = [e['date'] for e in new['weekend_shifts'] if not e['night']]
    print(f"  every weekend day has a night (MYT) worker: {'PASS' if not nights else 'FAIL '+str(nights)}")
    ok &= not nights
    print("  per-staff weekday cell coverage (Mon-Fri non-null, expect 22):")
    for s_name in new['staff_order']:
        cnt = 0
        for w in new['weeks']:
            st = next((x for x in w['staff'] if x['name']==s_name), None)
            if st:
                cnt += sum(1 for v in st['shifts'][:5] if v is not None)
        flag = '' if cnt == 22 else '   <-- CHECK'
        print(f"     {s_name:12} {cnt}{flag}")
        ok &= cnt == 22
    return ok

def patch_html(fp, data_json_str, ts):
    html = open(fp, encoding='utf-8').read()
    if 'data-month="Sep26"' not in html:
        html, c = re.subn(
            r'(<button class="month-btn)( active)?(" data-month="Aug26" onclick="switchMonth\(\'Aug26\'\)">Aug 2026</button>)',
            lambda m: m.group(1) + m.group(3) + '\n    <button class=\"month-btn active\" data-month=\"Sep26\" onclick=\"switchMonth(\'Sep26\')\">Sep 2026</button>',
            html, count=1)
        assert c == 1, f"{fp}: month button not inserted"
    if "var currentMonth = 'Sep26';" not in html:
        html, c = re.subn(r"var currentMonth = 'Aug26';", "var currentMonth = 'Sep26';", html)
        assert c == 1, f"{fp}: currentMonth not updated"
    if "monthKey === 'Sep26'" not in html:
        html, c = re.subn(
            r"(  if \(monthKey === 'Aug26'\))",
            "  if (monthKey === 'Sep26') return { year: 2026, month: 8, name: 'September 2026', days: 30 };\n\\1",
            html, count=1)
        assert c == 1, f"{fp}: getMonthInfo not updated"
    html, c = re.subn(r'var EMBEDDED_DATA = \{.*?\n\};',
                      'var EMBEDDED_DATA = ' + data_json_str + ';', html, count=1, flags=re.DOTALL)
    assert c == 1, f"{fp}: EMBEDDED_DATA not replaced"
    html, c = re.subn(r"var BUILD_TS = '\d+';", f"var BUILD_TS = '{ts}';", html)
    assert c >= 1, f"{fp}: BUILD_TS not replaced"
    open(fp, 'w', encoding='utf-8').write(html)
    print(f"{os.path.basename(fp)} patched (button + currentMonth + getMonthInfo + data + BUILD_TS={ts})")

if __name__ == '__main__':
    new, problems = extract()
    if problems:
        print("!!! WEEKEND ROLE PROBLEMS:")
        for p in problems: print("   ", p)
        print()
    print(json.dumps(new['weekend_shifts'], indent=1))
    ok = validate(new)

    if '--write' in sys.argv:
        if problems or not ok:
            raise SystemExit("Refusing to write: resolve problems above first.")
        data = json.load(open(DJ, encoding='utf-8'))
        if 'Sep26' in data:
            print("NOTE: Sep26 already in data.json - replacing it.")
        data['Sep26'] = new
        json.dump(data, open(DJ, 'w', encoding='utf-8'), indent=2, ensure_ascii=False)
        print("\ndata.json written (Sep26 added; other months untouched).")
        ts = str(int(time.time()))
        open(os.path.join(REPO, 'v.txt'), 'w').write(ts)
        print("v.txt:", ts)
        data_json_str = json.dumps(data, indent=2, ensure_ascii=False)
        for hf in ['index.html', 'management.html']:
            patch_html(os.path.join(REPO, hf), data_json_str, ts)
        print("\n=== WRITE COMPLETE ===")
